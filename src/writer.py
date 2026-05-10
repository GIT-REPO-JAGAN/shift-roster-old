"""
writer.py
---------
Generates the formatted Excel output workbook.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import AppConfig
from .scheduler import EmployeeSchedule, SHIFT_LABELS, iter_dates, is_weekend


# ─── Palette ───────────────────────────────────────────────────────────────────

CLR = {
    "hdr_bg":   "1F3864",
    "hdr_fg":   "FFFFFF",
    "sub_bg":   "2E75B6",
    "skill_bg": "D6E4F7",
    "alt_row":  "EBF3FB",
    "white":    "FFFFFF",
    "M":        "FFF2CC",
    "A":        "DDEBF7",
    "N":        "E2EFDA",
    "E":        "FCE4D6",
    "H":        "F4CCCC",
    "L":        "D9D2E9",
    "W":        "D9D9D9",
    "weekend":  "F2F2F2",
    "holiday":  "FFE6E6",
}


# ─── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(size=9, bold=False, color="000000", italic=False) -> Font:
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)


def _border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _align(horizontal="center", wrap=True) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


def _hdr_cell(ws, row, col, value, bg=None, fg=None, size=10, bold=True, align="center"):
    c = ws.cell(row, col, value)
    c.font      = _font(size=size, bold=bold, color=fg or CLR["hdr_fg"])
    c.fill      = _fill(bg or CLR["hdr_bg"])
    c.alignment = _align(align)
    c.border    = _border()
    return c


def _data_cell(ws, row, col, value, bg="FFFFFF", bold=False, align="center"):
    c = ws.cell(row, col, value)
    c.font      = _font(bold=bold)
    c.fill      = _fill(bg)
    c.alignment = _align(align)
    c.border    = _border()
    return c


# ─── Sheet writers ─────────────────────────────────────────────────────────────

def _write_roster_sheet(
    wb: Workbook,
    cfg: AppConfig,
    schedules: list[EmployeeSchedule],
) -> None:
    ws = wb.active
    ws.title = "Shift Roster"

    all_dates = list(iter_dates(cfg.roster_start, cfg.roster_end))
    total_cols = 4 + len(all_dates)
    hol_days = ", ".join(str(d.day) for d in sorted(cfg.account_holidays))

    # Row 1 – Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(1, 1, f"Shift Roster — {cfg.month_label}")
    c.font      = _font(size=14, bold=True, color=CLR["hdr_fg"])
    c.fill      = _fill(CLR["hdr_bg"])
    c.alignment = _align()
    ws.row_dimensions[1].height = 28

    # Row 2 – Legend hint
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    month_str = cfg.roster_start.strftime("%B %Y")
    c = ws.cell(
        2, 1,
        f"Account Holidays: {hol_days} {month_str}   |   "
        "M=Morning  A=Afternoon  N=Night  E=Evening  H=Holiday  L=Leave  W=WeekOff",
    )
    c.font      = _font(size=9, italic=True, color=CLR["hdr_fg"])
    c.fill      = _fill(CLR["sub_bg"])
    c.alignment = _align("left")
    ws.row_dimensions[2].height = 18

    # Row 3 – Column headers
    for ci, label in enumerate(["Name", "Email", "Skill", "Location"], start=1):
        _hdr_cell(ws, 3, ci, label)

    for di, d in enumerate(all_dates):
        ci = 5 + di
        label = f"{d.day}\n{d.strftime('%a')}"
        if d in cfg.account_holidays:
            _hdr_cell(ws, 3, ci, label, bg=CLR["holiday"], fg="CC0000", size=8)
        elif is_weekend(d):
            _hdr_cell(ws, 3, ci, label, bg=CLR["weekend"], fg="444444", size=8)
        else:
            _hdr_cell(ws, 3, ci, label, size=8)

    ws.row_dimensions[3].height = 30

    # Rows 4+ – Employee data grouped by skill
    row = 4
    skills_order = list(cfg.skill_rules.keys())

    # Index schedules by skill
    by_skill: dict[str, list[EmployeeSchedule]] = {s: [] for s in skills_order}
    for sched in schedules:
        if sched.employee.skill in by_skill:
            by_skill[sched.employee.skill].append(sched)

    for skill in skills_order:
        group = by_skill[skill]
        if not group:
            continue

        # Skill group header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        c = ws.cell(row, 1, f"  {skill}")
        c.font      = _font(size=10, bold=True, color=CLR["hdr_bg"])
        c.fill      = _fill(CLR["skill_bg"])
        c.alignment = _align("left")
        c.border    = _border()
        ws.row_dimensions[row].height = 20
        row += 1

        for emp_idx, sched in enumerate(group):
            emp = sched.employee
            bg  = CLR["white"] if emp_idx % 2 == 0 else CLR["alt_row"]

            for ci, val in enumerate(
                [emp.name, emp.email, emp.skill, emp.location], start=1
            ):
                _data_cell(ws, row, ci, val, bg=bg, align="left")

            for di, d in enumerate(all_dates):
                code = sched.daily.get(d, "")
                _data_cell(
                    ws, row, 5 + di, code,
                    bg=CLR.get(code, CLR["white"]),
                    bold=(code in ("H", "L", "W")),
                )

            ws.row_dimensions[row].height = 18
            row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    for di in range(len(all_dates)):
        ws.column_dimensions[get_column_letter(5 + di)].width = 4.5

    ws.freeze_panes = "E4"


def _write_summary_sheet(
    wb: Workbook,
    cfg: AppConfig,
    schedules: list[EmployeeSchedule],
) -> None:
    ws = wb.create_sheet("Summary")
    all_dates = list(iter_dates(cfg.roster_start, cfg.roster_end))
    headers = [
        "Name", "Skill", "Morning (M)", "Afternoon (A)", "Night (N)",
        "Evening (E)", "Week Off (W)", "Holiday (H)", "Leave (L)", "Working Days",
    ]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = ws.cell(1, 1, f"Shift Distribution Summary — {cfg.month_label}")
    c.font = _font(size=13, bold=True, color=CLR["hdr_fg"])
    c.fill = _fill(CLR["hdr_bg"])
    c.alignment = _align()
    ws.row_dimensions[1].height = 26

    for ci, h in enumerate(headers, 1):
        _hdr_cell(ws, 2, ci, h, bg=CLR["sub_bg"], size=9)
    ws.row_dimensions[2].height = 22

    for ri, sched in enumerate(schedules, start=3):
        counts = sched.shift_counts()
        emp    = sched.employee
        bg     = CLR["white"] if ri % 2 == 0 else CLR["alt_row"]
        vals   = [
            emp.name, emp.skill,
            counts.get("M", 0), counts.get("A", 0), counts.get("N", 0),
            counts.get("E", 0), counts.get("W", 0), counts.get("H", 0),
            counts.get("L", 0), sched.working_days(),
        ]
        for ci, v in enumerate(vals, 1):
            _data_cell(ws, ri, ci, v, bg=bg, align="left" if ci <= 2 else "center")
        ws.row_dimensions[ri].height = 16

    for ci, w in enumerate([28, 24, 13, 13, 13, 13, 13, 13, 13, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _write_legend_sheet(wb: Workbook, cfg: AppConfig) -> None:
    ws = wb.create_sheet("Legend")

    ws.merge_cells("A1:C1")
    c = ws.cell(1, 1, "Shift Codes & Colour Legend")
    c.font = _font(size=12, bold=True, color=CLR["hdr_fg"])
    c.fill = _fill(CLR["hdr_bg"])
    c.alignment = _align()
    ws.row_dimensions[1].height = 24

    for ci, h in enumerate(["Code", "Description", "Colour"], 1):
        _hdr_cell(ws, 2, ci, h, bg=CLR["sub_bg"])
    ws.row_dimensions[2].height = 20

    legend_rows = [
        ("M", SHIFT_LABELS["M"]),
        ("A", SHIFT_LABELS["A"]),
        ("N", SHIFT_LABELS["N"]),
        ("E", SHIFT_LABELS["E"]),
        ("H", SHIFT_LABELS["H"]),
        ("L", SHIFT_LABELS["L"]),
        ("W", SHIFT_LABELS["W"]),
    ]
    for ri, (code, desc) in enumerate(legend_rows, start=3):
        color = CLR.get(code, CLR["white"])
        c = ws.cell(ri, 1, code)
        c.font = _font(size=10, bold=True)
        c.fill = _fill(color)
        c.alignment = _align()
        c.border = _border()

        c = ws.cell(ri, 2, desc)
        c.font = _font(size=10)
        c.fill = _fill(CLR["white"])
        c.alignment = _align("left")
        c.border = _border()

        c = ws.cell(ri, 3)
        c.fill = _fill(color)
        c.border = _border()
        ws.row_dimensions[ri].height = 20

    # Week-off rules note
    note_row = len(legend_rows) + 4
    ws.merge_cells(f"A{note_row}:C{note_row}")
    c = ws.cell(note_row, 1, "Week-Off Rules")
    c.font = _font(size=10, bold=True, color=CLR["hdr_bg"])
    c.alignment = _align("left")

    weekend_skills = [s for s, r in cfg.skill_rules.items() if r.get("week_off") == "weekends"]
    every5_skills  = [s for s, r in cfg.skill_rules.items() if r.get("week_off") == "every5th"]

    rules = [
        f"Weekends (Sat & Sun): {', '.join(weekend_skills)}",
        f"Every 5th Day (rotating): {', '.join(every5_skills)}",
    ]
    for i, rule in enumerate(rules, start=note_row + 1):
        ws.merge_cells(f"A{i}:C{i}")
        c = ws.cell(i, 1, rule)
        c.font = _font(size=9)
        c.alignment = _align("left")
        ws.row_dimensions[i].height = 16

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16


# ─── Public API ────────────────────────────────────────────────────────────────

def write_workbook(
    cfg: AppConfig,
    schedules: list[EmployeeSchedule],
) -> Path:
    """
    Build and save the output workbook.
    Returns the path it was saved to.
    """
    cfg.output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _write_roster_sheet(wb, cfg, schedules)
    _write_summary_sheet(wb, cfg, schedules)
    _write_legend_sheet(wb, cfg)

    wb.save(cfg.output_file)
    return cfg.output_file
