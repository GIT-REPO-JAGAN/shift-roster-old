"""
writer.py
---------
Generates the formatted Excel output workbook.
Supports PL, CO, ADHOC day codes and multi-month rosters.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import AppConfig
from .scheduler import EmployeeSchedule, SHIFT_LABELS, iter_dates, is_weekend


# ─── Palette ───────────────────────────────────────────────────────────────────

CLR = {
    "hdr_bg":   "1F3864",
    "hdr_fg":   "FFFFFF",
    "sub_bg":   "2E75B6",
    "skill_bg": "D6E4F7",
    "alt_row":  "EBF3FB",
    "white":    "FFFFFF",
    "M":        "FFF2CC",
    "A":        "DDEBF7",
    "N":        "E2EFDA",
    "E":        "FCE4D6",
    "E1":       "F8D7F0",
    "G":        "FFF2CC",
    "PL":       "D9D2E9",
    "CO":       "C6EFCE",
    "ADHOC":    "FFE0CC",
    "H":        "F4CCCC",
    "W":        "D9D9D9",
    "weekend":  "F2F2F2",
    "holiday":  "FFE6E6",
}

NON_WORKING = {"H", "W", "PL", "CO"}  # ADHOC counts as a working day


# ─── Helpers ───────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(size=9, bold=False, color="000000", italic=False) -> Font:
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)


def _border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _align(horizontal="center", wrap=True) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


def _hdr_cell(ws, row, col, value, bg=None, fg=None, size=10, bold=True, align="center"):
    c = ws.cell(row, col, value)
    c.font      = _font(size=size, bold=bold, color=fg or CLR["hdr_fg"])
    c.fill      = _fill(bg or CLR["hdr_bg"])
    c.alignment = _align(align)
    c.border    = _border()
    return c


def _data_cell(ws, row, col, value, bg="FFFFFF", bold=False, align="center"):
    c = ws.cell(row, col, value)
    c.font      = _font(bold=bold)
    c.fill      = _fill(bg)
    c.alignment = _align(align)
    c.border    = _border()
    return c


def _period_label(start: date, end: date) -> str:
    """Return a human-readable period label covering the full date range."""
    if start.year == end.year and start.month == end.month:
        return start.strftime("%B %Y")
    if start.year == end.year:
        return f"{start.strftime('%B')} – {end.strftime('%B %Y')}"
    return f"{start.strftime('%B %Y')} – {end.strftime('%B %Y')}"


def _holidays_label(holidays: set[date], start: date, end: date) -> str:
    """Group holiday dates by month for the legend row."""
    from itertools import groupby
    sorted_hols = sorted(holidays)
    groups = {}
    for h in sorted_hols:
        key = h.strftime("%B %Y")
        groups.setdefault(key, []).append(str(h.day))
    return "  |  ".join(f"{', '.join(days)} {month}" for month, days in groups.items())


# ─── Roster sheet ──────────────────────────────────────────────────────────────

def _write_roster_sheet(
    wb: Workbook,
    cfg: AppConfig,
    schedules: list[EmployeeSchedule],
) -> None:
    ws = wb.active
    ws.title = "Shift Roster"

    all_dates  = list(iter_dates(cfg.roster_start, cfg.roster_end))
    total_cols = 5 + len(all_dates)

    period_lbl = _period_label(cfg.roster_start, cfg.roster_end)
    hol_lbl    = _holidays_label(cfg.account_holidays, cfg.roster_start, cfg.roster_end)
    if not hol_lbl:
        hol_lbl = "None"

    # Row 1 – Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(1, 1, f"Shift Roster — {period_lbl}")
    c.font      = _font(size=14, bold=True, color=CLR["hdr_fg"])
    c.fill      = _fill(CLR["hdr_bg"])
    c.alignment = _align()
    ws.row_dimensions[1].height = 28

    # Row 2 – Legend
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    # Build location holidays label for legend
    loc_hol_map = getattr(cfg, "location_holidays", {}) or {}
    loc_hol_parts = []
    for loc, dates in sorted(loc_hol_map.items()):
        loc_days = _holidays_label(dates, cfg.roster_start, cfg.roster_end)
        if loc_days:
            loc_hol_parts.append(f"{loc}: {loc_days}")
    loc_hol_lbl = "  |  ".join(loc_hol_parts) if loc_hol_parts else "None"

    legend_text = (
        f"AH: {hol_lbl}   |   LH: {loc_hol_lbl}   |   "
        "M=Morning  A=Afternoon  N=Night  E=Evening  E1=Evening-1  G=General  "
        "H=Holiday  PL=Planned Leave  CO=Comp-Off  ADHOC=Adhoc  W=Week Off"
    )
    c = ws.cell(2, 1, legend_text)
    c.font      = _font(size=9, italic=True, color=CLR["hdr_fg"])
    c.fill      = _fill(CLR["sub_bg"])
    c.alignment = _align("left")
    ws.row_dimensions[2].height = 18

    # Row 3 – Column headers
    for ci, label in enumerate(["Name", "Email", "Skill", "Location", "Shift"], 1):
        _hdr_cell(ws, 3, ci, label)

    for di, d in enumerate(all_dates):
        ci    = 6 + di
        label = f"{d.day}\n{d.strftime('%a')}"
        if d in cfg.account_holidays:
            _hdr_cell(ws, 3, ci, label, bg=CLR["holiday"], fg="CC0000", size=8)
        elif is_weekend(d):
            _hdr_cell(ws, 3, ci, label, bg=CLR["weekend"], fg="444444", size=8)
        else:
            _hdr_cell(ws, 3, ci, label, size=8)

    ws.row_dimensions[3].height = 30

    # Group by skill
    skills_order = list(dict.fromkeys(s.employee.skill for s in schedules))
    by_skill: dict[str, list[EmployeeSchedule]] = {}
    for sched in schedules:
        by_skill.setdefault(sched.employee.skill, []).append(sched)

    row = 4
    for skill in skills_order:
        group = by_skill.get(skill, [])
        if not group:
            continue

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        c = ws.cell(row, 1, f"  {skill}")
        c.font      = _font(size=10, bold=True, color=CLR["hdr_bg"])
        c.fill      = _fill(CLR["skill_bg"])
        c.alignment = _align("left")
        c.border    = _border()
        ws.row_dimensions[row].height = 20
        row += 1

        for emp_idx, sched in enumerate(group):
            emp = sched.employee
            bg  = CLR["white"] if emp_idx % 2 == 0 else CLR["alt_row"]

            primary_shift = next(
                (v for v in sched.daily.values() if v not in NON_WORKING and v != "ADHOC"),
                "—"
            )

            vals = [emp.name, emp.email, emp.skill, emp.location, primary_shift]
            for ci, val in enumerate(vals, 1):
                _data_cell(ws, row, ci, val, bg=bg, align="left" if ci <= 4 else "center")

            for di, d in enumerate(all_dates):
                ci   = 6 + di
                code = sched.daily.get(d, "")
                _data_cell(
                    ws, row, ci, code,
                    bg=CLR.get(code, CLR.get(code.rstrip("01"), CLR["white"])),
                    bold=(code in ("H", "PL", "CO", "W", "ADHOC")),
                )

            ws.row_dimensions[row].height = 18
            row += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 8
    for di in range(len(all_dates)):
        ws.column_dimensions[get_column_letter(6 + di)].width = 4.2

    ws.freeze_panes = "F4"


# ─── Summary sheet ─────────────────────────────────────────────────────────────

def _write_summary_sheet(
    wb: Workbook,
    cfg: AppConfig,
    schedules: list[EmployeeSchedule],
) -> None:
    ws = wb.create_sheet("Summary")
    period_lbl = _period_label(cfg.roster_start, cfg.roster_end)
    headers = [
        "Name", "Skill", "Shift",
        "Morning(M)", "Afternoon(A)", "Night(N)", "Evening(E)", "General(G)",
        "Week Off(W)", "Holiday(H)", "PL", "CO", "ADHOC", "Working Days",
    ]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = ws.cell(1, 1, f"Shift Distribution Summary — {period_lbl}")
    c.font      = _font(size=13, bold=True, color=CLR["hdr_fg"])
    c.fill      = _fill(CLR["hdr_bg"])
    c.alignment = _align()
    ws.row_dimensions[1].height = 26

    for ci, h in enumerate(headers, 1):
        _hdr_cell(ws, 2, ci, h, bg=CLR["sub_bg"], size=9)
    ws.row_dimensions[2].height = 22

    for ri, sched in enumerate(schedules, 3):
        counts  = sched.shift_counts()
        emp     = sched.employee
        bg      = CLR["white"] if ri % 2 == 0 else CLR["alt_row"]
        primary = next(
            (v for v in sched.daily.values() if v not in NON_WORKING and v != "ADHOC"), "—"
        )
        vals = [
            emp.name, emp.skill, primary,
            counts.get("M", 0), counts.get("A", 0),
            counts.get("N", 0), counts.get("E", 0), counts.get("G", 0),
            counts.get("W", 0), counts.get("H", 0),
            counts.get("PL", 0), counts.get("CO", 0), counts.get("ADHOC", 0),
            sched.working_days(),
        ]
        for ci, v in enumerate(vals, 1):
            _data_cell(ws, ri, ci, v, bg=bg, align="left" if ci <= 2 else "center")
        ws.row_dimensions[ri].height = 16

    widths = [26, 22, 7, 10, 11, 9, 9, 9, 10, 10, 7, 7, 8, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ─── Legend sheet ──────────────────────────────────────────────────────────────

def _write_legend_sheet(wb: Workbook, cfg: AppConfig) -> None:
    ws = wb.create_sheet("Legend")

    ws.merge_cells("A1:C1")
    c = ws.cell(1, 1, "Shift Codes & Colour Legend")
    c.font      = _font(size=12, bold=True, color=CLR["hdr_fg"])
    c.fill      = _fill(CLR["hdr_bg"])
    c.alignment = _align()
    ws.row_dimensions[1].height = 24

    for ci, h in enumerate(["Code", "Description", "Colour"], 1):
        _hdr_cell(ws, 2, ci, h, bg=CLR["sub_bg"])
    ws.row_dimensions[2].height = 20

    legend_rows = [
        ("M",     SHIFT_LABELS["M"]),
        ("A",     SHIFT_LABELS["A"]),
        ("N",     SHIFT_LABELS["N"]),
        ("E",     SHIFT_LABELS["E"]),
        ("G",     SHIFT_LABELS["G"]),
        ("PL",    SHIFT_LABELS["PL"]),
        ("CO",    SHIFT_LABELS["CO"]),
        ("ADHOC", "Adhoc / On-call Shift"),
        ("H",     SHIFT_LABELS["H"]),
        ("W",     SHIFT_LABELS["W"]),
    ]
    for ri, (code, desc) in enumerate(legend_rows, 3):
        color = CLR.get(code, CLR["white"])
        c = ws.cell(ri, 1, code)
        c.font      = _font(size=10, bold=True)
        c.fill      = _fill(color)
        c.alignment = _align()
        c.border    = _border()
        ws.cell(ri, 2, desc).font      = _font(size=10)
        ws.cell(ri, 2).alignment       = _align("left")
        ws.cell(ri, 2).border          = _border()
        ws.cell(ri, 3).fill            = _fill(color)
        ws.cell(ri, 3).border          = _border()
        ws.row_dimensions[ri].height   = 20

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 16

    note_row = len(legend_rows) + 4
    ws.merge_cells(f"A{note_row}:C{note_row}")
    ws.cell(note_row, 1, "Week-Off Rules").font = _font(size=10, bold=True, color=CLR["hdr_bg"])
    ws.cell(note_row, 1).alignment = _align("left")

    weekend_skills = [s for s, r in cfg.skill_rules.items() if r.get("week_off") == "weekends"]
    rolling_skills = [s for s, r in cfg.skill_rules.items() if "rolling" in r.get("week_off", "")]
    other_skills   = [s for s, r in cfg.skill_rules.items()
                      if r.get("week_off") not in ("weekends",)
                      and "rolling" not in r.get("week_off", "")]

    notes = []
    if weekend_skills:
        notes.append(f"Weekends (Sat & Sun): {', '.join(weekend_skills)}")
    if rolling_skills:
        notes.append(f"Rolling 6th & 7th day: {', '.join(rolling_skills)}")
    if other_skills:
        notes.append(f"Custom cycle: {', '.join(other_skills)}")

    for i, note in enumerate(notes, note_row + 1):
        ws.merge_cells(f"A{i}:C{i}")
        c = ws.cell(i, 1, note)
        c.font      = _font(size=9)
        c.alignment = _align("left")
        ws.row_dimensions[i].height = 16


# ─── Public API ────────────────────────────────────────────────────────────────

def write_workbook(cfg: AppConfig, schedules: list[EmployeeSchedule]) -> Path:
    cfg.output_file.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _write_roster_sheet(wb, cfg, schedules)
    _write_summary_sheet(wb, cfg, schedules)
    _write_legend_sheet(wb, cfg)
    wb.save(cfg.output_file)
    return cfg.output_file
